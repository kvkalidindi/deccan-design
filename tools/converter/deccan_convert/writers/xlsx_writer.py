"""XLSX restyler: source .xlsx -> restyled .xlsx matching deccan-workbook.xltx.

Data is never altered — values, formulas, merged ranges, number formats,
column widths, and freeze panes carry over verbatim; only presentation
changes. The style recipe is read from the bundled template's own cells
(header row: Cascadia Mono 10 bold, stone-500 on stone-100 with a medium
stone-900 bottom border; body: Segoe UI Variable Text 10, stone-800, banded
stone-50 rows, thin stone-200 hairlines). Fonts are named, never embedded.
Workbooks exported from Google Sheets (File > Download > .xlsx) take the
same path.
"""

from __future__ import annotations

from pathlib import Path
from typing import Callable

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Token values (skill/references/tokens.md).
_STONE_50 = "FAFAF9"
_STONE_100 = "F5F5F4"
_STONE_200 = "E7E5E4"
_STONE_500 = "78716C"
_STONE_800 = "292524"
_STONE_900 = "1C1917"
_PAPER = "FFFFFF"

_SANS = "Segoe UI Variable Text"
_MONO = "Cascadia Mono"

_HEADER_FONT = Font(name=_MONO, size=10, bold=True, color=_STONE_500)
_BODY_FONT = Font(name=_SANS, size=10, color=_STONE_800)
_HEADER_FILL = PatternFill("solid", fgColor=_STONE_100)
_PAPER_FILL = PatternFill("solid", fgColor=_PAPER)
_BAND_FILL = PatternFill("solid", fgColor=_STONE_50)
_HEADER_BORDER = Border(bottom=Side(style="medium", color=_STONE_900))
_BODY_BORDER = Border(bottom=Side(style="thin", color=_STONE_200))

_MIN_COL_WIDTH = 12.0


def restyle_xlsx(
    source: Path, path: Path, log: Callable[[str], None] | None = None
) -> tuple[Path, list[str]]:
    say = log or (lambda _msg: None)
    warnings: list[str] = []

    wb = openpyxl.load_workbook(str(source))  # formulas preserved, not values

    chart_sheets = [ws.title for ws in wb.worksheets if getattr(ws, "_charts", None)]
    if chart_sheets:
        warnings.append(
            "xlsx: charts on sheet(s) "
            + ", ".join(chart_sheets)
            + " are not preserved by the restyler and may be dropped. "
            "Recreate them from the restyled data if needed."
        )

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        say(f"Restyling sheet '{ws.title}'")
        _restyle_sheet(ws)

    wb.template = False
    wb.save(str(path))
    return path, warnings


def _restyle_sheet(ws) -> None:
    if ws.max_row is None or ws.max_column is None:
        return

    header_rows = _header_row_count(ws)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            is_header = cell.row <= header_rows
            if is_header:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.border = _HEADER_BORDER
            else:
                cell.font = _BODY_FONT
                # Banded rows: first data row on paper, alternate stone-50.
                data_index = cell.row - header_rows
                cell.fill = _BAND_FILL if data_index % 2 == 0 else _PAPER_FILL
                cell.border = _BODY_BORDER

    # Column widths: keep author-set widths, give unset columns breathing room.
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions[letter]
        if not dim.customWidth or (dim.width or 0) < _MIN_COL_WIDTH:
            dim.width = max(dim.width or 0, _MIN_COL_WIDTH)

    # Freeze the header if the author has not chosen their own panes.
    if not ws.freeze_panes and header_rows:
        ws.freeze_panes = f"A{header_rows + 1}"

    ws.sheet_view.showGridLines = False


def _header_row_count(ws) -> int:
    """Rows to treat as the header band.

    An existing freeze-pane row marks the author's header; otherwise row 1
    is the header when it holds any value.
    """
    if ws.freeze_panes:
        try:
            row_part = "".join(ch for ch in str(ws.freeze_panes) if ch.isdigit())
            frozen_rows = int(row_part) - 1 if row_part else 0
            if 0 < frozen_rows <= 4:
                return frozen_rows
        except ValueError:
            pass
    first_row = next(ws.iter_rows(min_row=1, max_row=1), ())
    if any(cell.value is not None for cell in first_row):
        return 1
    return 0

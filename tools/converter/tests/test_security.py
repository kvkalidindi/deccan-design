"""Security regression tests for the converter's untrusted-input handling.

Each test pins a fix from the security audit: HTML sanitization of
render-time resource loads, metadata injection escaping, and the
resource-exhaustion guards (zip bombs, entity DTDs, declared-dimension OOM,
oversized data URIs, input-size and page caps, and the overwrite guard).
"""

import io
import zipfile
from pathlib import Path

import pytest

from deccan_convert import limits
from deccan_convert.convert import convert
from deccan_convert.ir import DocumentIR, Metadata
from deccan_convert.limits import InputTooLarge
from deccan_convert.readers._sections import build_sections
from deccan_convert.writers.html_writer import render_html


class TestSanitizerBlocksRenderExfiltration:
    def test_remote_img_dropped(self):
        out = build_sections('<h1>T</h1><p><img src="http://attacker/beacon.png?x=1"></p>')
        assert "attacker" not in out
        assert "beacon" not in out

    def test_file_img_dropped(self):
        out = build_sections('<h1>T</h1><p><img src="file:///etc/passwd"></p>')
        assert "etc/passwd" not in out

    def test_protocol_relative_img_dropped(self):
        out = build_sections('<h1>T</h1><p><img src="//attacker/x.png"></p>')
        assert "attacker" not in out

    def test_data_image_preserved(self):
        px = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAAAAAA6fptVAAAACklEQVR4nGP4DwABBAEAX+RA9wAAAABJRU5ErkJggg=="
        out = build_sections(f'<h1>T</h1><p><img src="{px}"></p>')
        assert "data:image/png;base64," in out

    def test_javascript_href_dropped(self):
        out = build_sections('<h1>T</h1><p><a href="javascript:alert(1)">x</a></p>')
        assert "javascript" not in out

    def test_obfuscated_javascript_href_dropped(self):
        out = build_sections('<h1>T</h1><p><a href="java\tscript:alert(1)">x</a></p>')
        assert "javascript" not in out.replace("\t", "")
        assert "alert" not in out

    def test_data_html_href_dropped(self):
        out = build_sections('<h1>T</h1><p><a href="data:text/html,<script>1</script>">x</a></p>')
        assert "data:text/html" not in out

    def test_safe_href_preserved(self):
        out = build_sections('<h1>T</h1><p><a href="https://deccanchemicals.com">x</a></p>')
        assert "https://deccanchemicals.com" in out

    def test_svg_and_style_still_stripped(self):
        out = build_sections(
            '<h1>T</h1><svg><image href="http://attacker/x"/></svg>'
            '<div style="background:url(http://attacker/y)">z</div>'
        )
        assert "attacker" not in out
        assert "<svg" not in out.lower()


class TestMetadataInjectionEscaped:
    def test_title_cannot_break_out(self):
        ir = DocumentIR(
            metadata=Metadata(
                title='</title><script>alert(1)</script>',
                document_type="Report",
                prepared_by="X",
            ),
            body_html="<section class='section'><h1>T</h1><p>b</p></section>",
        )
        html = render_html(ir)
        assert "<script>alert(1)</script>" not in html
        assert "&lt;script&gt;" in html


class TestResourceGuards:
    def _zip_with_entry(self, tmp_path: Path, name: str, size: int, member="x.xml") -> Path:
        path = tmp_path / name
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
            z.writestr(member, b"A" * size)
        return path

    def test_zip_bomb_ratio_rejected(self, tmp_path):
        bomb = self._zip_with_entry(tmp_path, "bomb.xlsx", 300 * 1024 * 1024)
        with pytest.raises(InputTooLarge):
            limits.guard_zip(bomb)

    def test_too_many_entries_rejected(self, tmp_path):
        path = tmp_path / "many.pptx"
        with zipfile.ZipFile(path, "w") as z:
            for i in range(limits.MAX_ZIP_ENTRIES + 1):
                z.writestr(f"e{i}.xml", b"x")
        with pytest.raises(InputTooLarge):
            limits.guard_zip(path)

    def test_oversized_input_rejected(self, tmp_path, monkeypatch):
        monkeypatch.setattr(limits, "MAX_INPUT_BYTES", 1024)
        big = tmp_path / "big.md"
        big.write_bytes(b"#" * 4096)
        with pytest.raises(InputTooLarge):
            limits.guard_input_size(big)

    def test_docx_doctype_rejected(self, tmp_path):
        # A minimal zip posing as docx with a DTD in document.xml.
        path = tmp_path / "evil.docx"
        payload = b'<?xml version="1.0"?><!DOCTYPE x [<!ENTITY a "b">]><document/>'
        with zipfile.ZipFile(path, "w") as z:
            z.writestr("word/document.xml", payload)
        with pytest.raises(InputTooLarge, match="entity"):
            limits.guard_no_doctype(path, "word/document.xml")

    def test_convert_rejects_oversized(self, tmp_path, monkeypatch):
        monkeypatch.setattr(limits, "MAX_INPUT_BYTES", 512)
        src = tmp_path / "in.md"
        src.write_text("# T\n\n" + "word " * 500, encoding="utf-8")
        with pytest.raises(InputTooLarge):
            convert(src, tmp_path / "out.html",
                    metadata=Metadata(title="T", document_type="Report", prepared_by="X"))

    def test_xlsx_declared_dimension_clamped(self, tmp_path, monkeypatch):
        import openpyxl

        from deccan_convert.writers.xlsx_writer import restyle_xlsx

        monkeypatch.setattr("deccan_convert.writers.xlsx_writer.MAX_XLSX_ROWS", 100)
        monkeypatch.setattr("deccan_convert.writers.xlsx_writer.MAX_XLSX_COLS", 10)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Header"
        ws["ZZ5000"] = "far"  # declares a large used range
        src = tmp_path / "wide.xlsx"
        wb.save(str(src))

        out, warnings = restyle_xlsx(src, tmp_path / "out.xlsx")
        assert any("capped" in w for w in warnings)
        # Data outside the styled region is still present.
        assert openpyxl.load_workbook(str(out)).active["ZZ5000"].value == "far"

    def test_oversized_data_uri_image_skipped(self, tmp_path, monkeypatch):
        import docx

        from deccan_convert.writers.docx_writer import write_docx

        monkeypatch.setattr("deccan_convert.writers.docx_writer.MAX_DATA_URI_B64", 100)
        huge = "data:image/png;base64," + ("A" * 5000)
        ir = DocumentIR(
            metadata=Metadata(title="T", document_type="Report", prepared_by="X"),
            body_html=f'<section class="section"><h1>S</h1><p><img src="{huge}"></p></section>',
        )
        out = write_docx(ir, tmp_path / "o.docx")
        names = zipfile.ZipFile(out).namelist()
        assert not any(n.startswith("word/media/") for n in names)
        assert any("could not be embedded" in w for w in ir.warnings)


class TestOverwriteGuard:
    def test_case_insensitive_same_file_rejected(self, tmp_path, monkeypatch):
        # Simulate a case-insensitive FS: os.path.samefile reports identity
        # for the differently-cased path.
        src = tmp_path / "Report.docx"
        src.write_bytes(b"PK\x03\x04")
        out = tmp_path / "report.docx"

        import deccan_convert.convert as conv

        monkeypatch.setattr(conv.os.path, "samefile", lambda a, b: True)
        # Make the output "exist" so the guard engages.
        out.write_bytes(b"x")
        with pytest.raises(ValueError, match="overwritten"):
            convert(src, out, metadata=Metadata(title="T", document_type="Report", prepared_by="X"))
